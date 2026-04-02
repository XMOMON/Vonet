import io
from datetime import datetime, timezone
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models.all import Trade, Signal, Position, SignalStatus, PositionStatus

router = APIRouter()

# ── Shared style helpers ──────────────────────────────────────────────────────
GOLD   = "FFD700"
BLACK  = "0B0E11"
GREEN  = "0ECB81"
RED    = "F6465D"
GRAY   = "1E2026"
WHITE  = "EAECEF"
MUTED  = "848E9C"

def _header_fill():  return PatternFill("solid", fgColor=BLACK)
def _gold_fill():    return PatternFill("solid", fgColor="1A1A12")
def _thin_border():
    s = Side(style="thin", color="2B2F36")
    return Border(left=s, right=s, top=s, bottom=s)

def _style_header_row(ws, row_num: int, ncols: int):
    fill = PatternFill("solid", fgColor="141414")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = Font(bold=True, color=GOLD, name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()

def _style_data_cell(cell, color=WHITE, bold=False, align="left"):
    cell.fill = PatternFill("solid", fgColor="0F0F0F")
    cell.font = Font(color=color, name="Calibri", size=10, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _thin_border()

def _col_widths(ws, widths: list):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _title_row(ws, title: str, ncols: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.fill = PatternFill("solid", fgColor="0B0B08")
    cell.font = Font(bold=True, color=GOLD, name="Calibri", size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32


# ── Trades Sheet ──────────────────────────────────────────────────────────────
def _build_trades_sheet(wb, trades):
    ws = wb.active
    ws.title = "Trades"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    headers = ["ID", "Pair", "Entry", "Exit", "PnL (USD)", "PnL (%)", "Exit Reason", "Opened At", "Closed At", "Duration (min)"]
    ncols = len(headers)
    _title_row(ws, "Pro Paper Trader — Trade History", ncols)
    ws.row_dimensions[2].height = 22
    for col, h in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=h)
    _style_header_row(ws, 2, ncols)

    for i, t in enumerate(trades, start=3):
        ws.row_dimensions[i].height = 20
        pnl = round(t.pnl_usd or 0, 2)
        pnl_pct = round(t.pnl_pct or 0, 4)
        duration = ""
        if t.opened_at and t.closed_at:
            duration = round((t.closed_at - t.opened_at).total_seconds() / 60, 1)

        values = [
            t.id, t.pair,
            round(t.entry or 0, 6), round(t.exit or 0, 6),
            pnl, pnl_pct,
            t.exit_reason or "",
            t.opened_at.strftime("%Y-%m-%d %H:%M:%S") if t.opened_at else "",
            t.closed_at.strftime("%Y-%m-%d %H:%M:%S") if t.closed_at else "",
            duration,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            # PnL columns get green/red colouring
            if col == 5:
                _style_data_cell(cell, color=GREEN if pnl >= 0 else RED, bold=True, align="right")
            elif col == 6:
                _style_data_cell(cell, color=GREEN if pnl_pct >= 0 else RED, align="right")
            elif col in (3, 4):
                _style_data_cell(cell, align="right")
            else:
                _style_data_cell(cell)

    # Totals row
    if trades:
        tot_row = len(trades) + 3
        ws.row_dimensions[tot_row].height = 22
        total_pnl = sum(round(t.pnl_usd or 0, 2) for t in trades)
        wins = sum(1 for t in trades if (t.pnl_usd or 0) > 0)
        ws.cell(row=tot_row, column=1, value="TOTALS")
        ws.cell(row=tot_row, column=5, value=round(total_pnl, 2))
        ws.cell(row=tot_row, column=6, value=f"{round(wins/len(trades)*100,1)}% win rate")
        for col in range(1, ncols + 1):
            cell = ws.cell(row=tot_row, column=col)
            cell.fill = PatternFill("solid", fgColor="1A1A12")
            cell.font = Font(bold=True, color=GOLD, name="Calibri", size=10)
            cell.border = _thin_border()

    _col_widths(ws, [8, 14, 14, 14, 12, 10, 20, 20, 20, 14])


# ── Signals Sheet ─────────────────────────────────────────────────────────────
def _build_signals_sheet(wb, signals):
    ws = wb.create_sheet("Signals")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    headers = ["ID", "Pair", "Direction", "Entry", "TP1", "TP2", "SL", "R:R", "Confidence", "Status", "Source", "Created At"]
    ncols = len(headers)
    _title_row(ws, "Pro Paper Trader — Signals", ncols)
    for col, h in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=h)
    _style_header_row(ws, 2, ncols)

    for i, s in enumerate(signals, start=3):
        ws.row_dimensions[i].height = 20
        entry = s.entry or 0
        sl = s.sl or 0
        tp2 = s.tp2 or 0
        direction = s.direction.value if hasattr(s.direction, "value") else str(s.direction)
        rr = ""
        if entry and sl and tp2 and (entry - sl) != 0:
            rr_val = ((tp2 - entry) / (entry - sl)) if direction == "LONG" else ((entry - tp2) / (sl - entry))
            rr = f"{round(rr_val, 2)}R"

        values = [
            s.id, s.pair, direction,
            round(entry, 6), round(s.tp1 or 0, 6), round(tp2, 6), round(sl, 6),
            rr,
            (s.confidence.value if hasattr(s.confidence, "value") else str(s.confidence or "")),
            (s.status.value if hasattr(s.status, "value") else str(s.status or "")),
            s.source or "",
            s.created_at.strftime("%Y-%m-%d %H:%M:%S") if s.created_at else "",
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            if col == 3:
                _style_data_cell(cell, color=GREEN if direction == "LONG" else RED, bold=True, align="center")
            else:
                _style_data_cell(cell)

    _col_widths(ws, [8, 14, 12, 14, 14, 14, 14, 8, 14, 12, 14, 20])


# ── Summary Sheet ─────────────────────────────────────────────────────────────
def _build_summary_sheet(wb, trades, signals):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    total = len(trades)
    wins = sum(1 for t in trades if (t.pnl_usd or 0) > 0)
    losses = total - wins
    total_pnl = sum(t.pnl_usd or 0 for t in trades)
    best = max((t.pnl_usd or 0 for t in trades), default=0)
    worst = min((t.pnl_usd or 0 for t in trades), default=0)
    win_rate = round(wins / total * 100, 1) if total else 0
    avg_pnl = round(total_pnl / total, 2) if total else 0

    ws.merge_cells("A1:D1")
    ws["A1"] = "Pro Paper Trader — Summary Report"
    ws["A1"].font = Font(bold=True, color=GOLD, name="Calibri", size=16)
    ws["A1"].fill = PatternFill("solid", fgColor="0B0B08")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A3"].font = Font(color=MUTED, name="Calibri", size=9)

    rows = [
        ("TRADING PERFORMANCE", ""),
        ("Total Trades", total),
        ("Winning Trades", wins),
        ("Losing Trades", losses),
        ("Win Rate", f"{win_rate}%"),
        ("Net P&L (USD)", round(total_pnl, 2)),
        ("Average P&L", round(avg_pnl, 2)),
        ("Best Trade", round(best, 2)),
        ("Worst Trade", round(worst, 2)),
        ("", ""),
        ("SIGNALS", ""),
        ("Total Signals", len(signals)),
        ("Pending", sum(1 for s in signals if s.status == SignalStatus.PENDING)),
        ("Executed", sum(1 for s in signals if s.status == SignalStatus.EXECUTED)),
        ("Cancelled", sum(1 for s in signals if s.status and s.status.value == "CANCELLED")),
    ]

    for r, (label, val) in enumerate(rows, start=5):
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=val)
        is_section = label in ("TRADING PERFORMANCE", "SIGNALS", "")
        lc.font = Font(bold=is_section, color=GOLD if is_section else WHITE, name="Calibri", size=10)
        vc.font = Font(
            bold=True,
            color=(GREEN if isinstance(val, (int, float)) and val > 0 else RED if isinstance(val, (int, float)) and val < 0 else WHITE),
            name="Calibri", size=10
        )
        lc.fill = PatternFill("solid", fgColor="0F0F0F")
        vc.fill = PatternFill("solid", fgColor="0F0F0F")
        lc.border = _thin_border()
        vc.border = _thin_border()
        lc.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 20

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18


# ── Main endpoint ─────────────────────────────────────────────────────────────
@router.get("/trades", summary="Export all trades as Excel (.xlsx)")
async def export_trades_excel(db: AsyncSession = Depends(get_db)):
    """Download a styled Excel workbook with trades, signals, and summary tabs."""
    trades_result = await db.execute(select(Trade).order_by(Trade.closed_at.desc()))
    trades = trades_result.scalars().all()

    signals_result = await db.execute(select(Signal).order_by(Signal.created_at.desc()))
    signals = signals_result.scalars().all()

    wb = openpyxl.Workbook()
    _build_summary_sheet(wb, trades, signals)
    _build_trades_sheet(wb, trades)
    _build_signals_sheet(wb, signals)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"paper_trader_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
